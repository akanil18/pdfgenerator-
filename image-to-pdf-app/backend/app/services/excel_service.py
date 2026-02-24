"""
PDF to Excel Service — extracts tables from a PDF into an XLSX file.
"""

import os
import uuid
import logging

import pdfplumber
from openpyxl import Workbook

logger = logging.getLogger(__name__)


def pdf_to_excel(pdf_path: str, session_dir: str) -> str:
    """
    Extract tables from a PDF and write them to an Excel workbook.
    Each PDF page with a table becomes a sheet in the workbook.
    Returns the path to the generated XLSX file.
    """
    xlsx_filename = f"{uuid.uuid4().hex}.xlsx"
    xlsx_path = os.path.join(session_dir, xlsx_filename)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    tables_found = 0

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if not tables:
                    continue

                for t_idx, table in enumerate(tables):
                    tables_found += 1
                    sheet_name = f"Page{page_num}"
                    if t_idx > 0:
                        sheet_name += f"_T{t_idx + 1}"
                    # Excel sheet names max 31 chars
                    sheet_name = sheet_name[:31]
                    ws = wb.create_sheet(title=sheet_name)

                    for row in table:
                        cleaned = [cell if cell is not None else "" for cell in row]
                        ws.append(cleaned)

                    logger.info(f"Extracted table from page {page_num} (table {t_idx + 1})")

    except Exception as exc:
        logger.error(f"Failed to extract tables from PDF: {exc}")
        raise ValueError(f"PDF to Excel conversion failed: {exc}") from exc

    # If no tables found, create a sheet with a notice
    if tables_found == 0:
        ws = wb.create_sheet(title="Info")
        ws.append(["No tables were found in the PDF."])
        ws.append(["Try a PDF that contains tabular data."])
        logger.warning("No tables found in the uploaded PDF.")

    wb.save(xlsx_path)

    file_size = os.path.getsize(xlsx_path)
    logger.info(f"✅  XLSX created: {xlsx_path} ({file_size:,} bytes, {tables_found} tables)")
    return xlsx_path
